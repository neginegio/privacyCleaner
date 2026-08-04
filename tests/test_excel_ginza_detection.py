from __future__ import annotations

import sys
import tempfile
from pathlib import Path

from openpyxl import Workbook, load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from excel_privacy_cleaner.excel_processor import (  # noqa: E402
    EXCEL_NLP_DETECTION_KIND,
    ExcelPrivacyProcessor,
)


def assert_true(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def create_bare_company_name_fixture(path: Path) -> None:
    # Reproduces the same real-world recall gap fixed in Word: a company's
    # short form used without its legal-entity suffix (株式会社 etc.) is
    # invisible to the column-header rules and Presidio, since neither
    # requires the AI-only structural cue GiNZA uses.
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet["A1"] = "備考"
    sheet["A2"] = "場所　アルプススチール様1階応接室"
    workbook.save(path)


def test_ginza_catches_bare_company_name_pattern_rules_miss() -> None:
    with tempfile.TemporaryDirectory(prefix="excel_ginza_") as tmpdir:
        tmp = Path(tmpdir)
        source = tmp / "fixture.xlsx"
        create_bare_company_name_fixture(source)

        processor = ExcelPrivacyProcessor()
        findings = processor.scan(source)

        ginza_findings = [f for f in findings if f.detection_kind == EXCEL_NLP_DETECTION_KIND]
        assert_true(
            bool(ginza_findings),
            "GiNZA should surface a candidate for the bare company name that column rules/Presidio structurally cannot see",
        )
        assert_true(
            any(f.original == "アルプススチール" for f in ginza_findings),
            "GiNZA candidate should cover the bare company name",
        )
        for finding in ginza_findings:
            assert_true(not finding.enabled, "AI-sourced Excel findings must always default to review-required")
            assert_true("ginza_ner" in finding.reason, "Reason text must clearly identify this as an AI/NLP judgment")


def test_ai_candidate_blocks_conversion_until_reviewed() -> None:
    with tempfile.TemporaryDirectory(prefix="excel_ginza_") as tmpdir:
        tmp = Path(tmpdir)
        source = tmp / "fixture.xlsx"
        create_bare_company_name_fixture(source)

        processor = ExcelPrivacyProcessor()
        findings = processor.scan(source)

        raised = False
        try:
            processor.convert_with_artifacts(source, findings, output_dir=tmp, write_artifacts=False)
        except RuntimeError as exc:
            raised = True
            assert_true("未確認候補" in str(exc), "Error should identify the block as an unreviewed candidate")
        assert_true(raised, "Leaving an AI-sourced candidate disabled must block conversion")


def test_ai_candidate_converts_when_approved() -> None:
    with tempfile.TemporaryDirectory(prefix="excel_ginza_") as tmpdir:
        tmp = Path(tmpdir)
        source = tmp / "fixture.xlsx"
        create_bare_company_name_fixture(source)

        processor = ExcelPrivacyProcessor()
        findings = processor.scan(source)
        for finding in findings:
            if finding.detection_kind == EXCEL_NLP_DETECTION_KIND:
                finding.enabled = True

        result = processor.convert_with_artifacts(source, findings, output_dir=tmp, write_artifacts=False)

        output = load_workbook(result.excel_path)
        cell_text = output["Sheet1"]["A2"].value
        assert_true("アルプススチール" not in cell_text, "Approved AI candidate should be replaced")
