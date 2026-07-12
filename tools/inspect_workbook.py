from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook


def safe(value: object) -> str:
    if value is None:
        return ""
    return str(value).encode("unicode_escape").decode("ascii")


def main() -> int:
    if len(sys.argv) > 1:
        path = Path(sys.argv[1])
    else:
        path = next(Path.cwd().glob("*検証結果*.xlsx"))

    workbook = load_workbook(path, data_only=True)
    print("workbook", safe(path.name))
    print("sheets", [(safe(ws.title), ws.max_row, ws.max_column) for ws in workbook.worksheets])
    for sheet in workbook.worksheets:
        print("##", safe(sheet.title))
        for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 80), values_only=True):
            print([safe(value) for value in row])
    workbook.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
