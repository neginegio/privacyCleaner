from __future__ import annotations

from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


def safe(value: object) -> str:
    return str(value).encode("unicode_escape").decode("ascii") if value is not None else ""


def main() -> int:
    workbook = load_workbook("Excel匿名化アプリ_検証結果_20260712_020935.xlsx", data_only=True)
    sheet = workbook["要修正一覧"]
    headers = [cell.value for cell in sheet[1]]

    rows: list[dict[str, object]] = []
    groups: dict[str, list[dict[str, object]]] = defaultdict(list)
    for row_values in sheet.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, row_values))
        if row.get("No.") is None:
            continue
        rows.append(row)
        groups[str(row.get("問題分類"))].append(row)

    print("count", len(rows))
    for problem, items in groups.items():
        print("##", safe(problem), len(items))
        for item in items[:8]:
            print(
                safe(item.get("No.")),
                safe(item.get("重要度")),
                safe(item.get("シート")),
                safe(item.get("セル")),
                safe(item.get("元データ")),
                "=>",
                safe(item.get("期待する出力・動作")),
            )

    instructions = workbook["Codex修正指示"]
    print("##INSTRUCTIONS")
    for row_values in instructions.iter_rows(min_row=4, values_only=True):
        if row_values[0] is None:
            continue
        print(" | ".join(safe(value) for value in row_values[:5]))

    workbook.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
