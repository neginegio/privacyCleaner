from __future__ import annotations

import argparse
from collections import Counter
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from excel_privacy_cleaner.excel_processor import ExcelPrivacyProcessor, ProcessingOptions


def cell_text(value) -> str:
    if value is None:
        return ""
    return str(value)


def compare_workbooks(actual_path: Path, expected_path: Path, max_diffs: int) -> tuple[int, Counter[str], list[str]]:
    actual = load_workbook(actual_path, data_only=False)
    expected = load_workbook(expected_path, data_only=False)
    diffs: list[str] = []
    sheet_counts: Counter[str] = Counter()
    total = 0
    try:
        for sheet_name in expected.sheetnames:
            if sheet_name not in actual.sheetnames:
                total += 1
                sheet_counts[sheet_name] += 1
                if len(diffs) < max_diffs:
                    diffs.append(f"{sheet_name}: actual workbook has no sheet")
                continue
            actual_sheet = actual[sheet_name]
            expected_sheet = expected[sheet_name]
            for row in range(1, expected_sheet.max_row + 1):
                for col in range(1, expected_sheet.max_column + 1):
                    coordinate = expected_sheet.cell(row=row, column=col).coordinate
                    actual_value = cell_text(actual_sheet[coordinate].value)
                    expected_value = cell_text(expected_sheet[coordinate].value)
                    if actual_value == expected_value:
                        continue
                    total += 1
                    sheet_counts[sheet_name] += 1
                    if len(diffs) < max_diffs:
                        diffs.append(
                            f"{sheet_name}!{coordinate}: actual={actual_value!r} expected={expected_value!r}"
                        )
        return total, sheet_counts, diffs
    finally:
        actual.close()
        expected.close()


def main() -> int:
    parser = argparse.ArgumentParser(description="Run anonymization and compare with an expected Excel workbook.")
    parser.add_argument("source", type=Path, help="Original input workbook")
    parser.add_argument("expected", type=Path, help="Expected anonymized workbook")
    parser.add_argument("--mode", choices=("analysis", "external"), default="analysis")
    parser.add_argument("--approve-candidates", action="store_true", help="Treat review candidates as approved candidate masks")
    parser.add_argument("--max-diffs", type=int, default=50)
    args = parser.parse_args()

    options = ProcessingOptions(mode=args.mode, transform_business_secrets=args.mode == "external")
    processor = ExcelPrivacyProcessor()
    findings = processor.scan(args.source, options=options)
    candidates = [finding for finding in findings if finding.detection_kind == "確認候補"]
    blocked_candidates = [finding for finding in candidates if not finding.enabled]
    if blocked_candidates and not args.approve_candidates:
        print(f"BLOCKED_REVIEW_REQUIRED count={len(blocked_candidates)}")
        for finding in blocked_candidates[: args.max_diffs]:
            print(f"{finding.sheet}!{finding.cell}: {finding.original!r} -> {finding.replacement!r}")
        processor.cleanup()
        return 2

    if args.approve_candidates:
        for finding in candidates:
            finding.enabled = True

    with tempfile.TemporaryDirectory(prefix="ExcelPrivacyExpected_") as tmp:
        actual_path = processor.convert(args.source, findings, Path(tmp), options=options)
        total_diffs, sheet_counts, diffs = compare_workbooks(actual_path, args.expected, args.max_diffs)
        if diffs:
            print(f"DIFF total={total_diffs}")
            for sheet_name, count in sheet_counts.items():
                print(f"SHEET {sheet_name}: {count}")
            for diff in diffs:
                print(diff)
            return 1

    print("OK expected workbook matched")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
