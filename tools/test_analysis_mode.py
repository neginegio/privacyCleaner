from __future__ import annotations

import re
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from excel_privacy_cleaner.excel_processor import ExcelPrivacyProcessor, ProcessingOptions  # noqa: E402


SOURCE = Path("Excel匿名化アプリ_テストデータ.xlsx")


def assert_true(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def main() -> int:
    processor = ExcelPrivacyProcessor()
    options = ProcessingOptions(mode="analysis")
    findings = processor.scan(SOURCE, options=options)

    with tempfile.TemporaryDirectory(prefix="privacy_analysis_test_") as tmp:
        result = processor.convert_with_artifacts(SOURCE, findings, Path(tmp), options=options)
        source_workbook = load_workbook(SOURCE, data_only=False)
        workbook = load_workbook(result.excel_path, data_only=False)
        try:
            source_customer_sheet = source_workbook["01_顧客マスタ"]
            deal_sheet = workbook["04_取引機密"]
            customer_sheet = workbook["01_顧客マスタ"]
            hr_sheet = workbook["03_人事情報"]
            free_sheet = workbook["02_自由記述"]
            variants_sheet = workbook["05_表記ゆれ"]

            for row in range(2, 7):
                formula_cell = deal_sheet[f"G{row}"]
                assert_true(formula_cell.data_type == "f", f"G{row} should remain a formula")
                assert_true(formula_cell.value == f"=IFERROR((E{row}-F{row})/E{row},0)", f"G{row} formula changed")
                assert_true(isinstance(deal_sheet[f"E{row}"].value, (int, float)), f"E{row} should remain numeric")
                assert_true(isinstance(deal_sheet[f"F{row}"].value, (int, float)), f"F{row} should remain numeric")

            for row in range(2, 11):
                if hasattr(source_customer_sheet[f"L{row}"].value, "year"):
                    assert_true(hasattr(customer_sheet[f"L{row}"].value, "year"), f"L{row} should remain date-like")

            assert_true(customer_sheet["B2"].value == customer_sheet["C2"].value, "Name and kana in the same row should share ID")
            assert_true(customer_sheet["B3"].value != hr_sheet["B4"].value, "Customer and employee Sato Hanako must not merge by name only")

            missing_words = {"匿名希望", "なし", "未回答", "不明", "該当なし"}
            for finding in findings:
                assert_true(finding.original not in missing_words, f"Missing-value word was detected: {finding.original}")

            review_originals = {(finding.sheet, finding.cell, finding.original) for finding in findings if finding.detection_kind == "確認候補"}
            assert_true(any(original == "森さん" for _sheet, _cell, original in review_originals), "森さん should be review candidate")
            assert_true(any(original == "南さん" for _sheet, _cell, original in review_originals), "南さん should be review candidate")

            assert_true("個人001です" in free_sheet["D2"].value, "Free-text full name should be partially replaced")
            assert_true("旧姓情報あり" in customer_sheet["O3"].value, "Old surname note should be generalized")
            assert_true("個人005" in free_sheet["D6"].value, "Romanized name should be replaced")

            secret_text = "\n".join(str(deal_sheet[cell].value) for cell in ("I2", "K2", "L2"))
            assert_true("1234567" not in secret_text, "Bank account number should not remain")
            assert_true("sk-test" not in secret_text, "API key should not remain")
            assert_true("123456789012" not in str(customer_sheet["N2"].value), "Personal number should not remain")

            assert_true("森さん" not in variants_sheet["C19"].value, "Review candidate original should not remain in output")
            assert_true("南さん" not in variants_sheet["C17"].value, "Review candidate original should not remain in output")
        finally:
            workbook.close()
            source_workbook.close()

        allowed_fixture_warning_fragments = (
            "xl/worksheets/sheet6.xml",
            "xl/worksheets/sheet7.xml",
            "xl/styles.xml",
        )
        leak_warnings = [
            warning
            for warning in result.warnings
            if "内部XML残存候補" in warning
            and not any(fragment in warning for fragment in allowed_fixture_warning_fragments)
        ]
        assert_true(not leak_warnings, f"Internal XML leak warnings found: {leak_warnings[:5]}")
        assert_true(result.formula_changed_count == 0, "Formula changed count should be zero")
        assert_true(result.csv_path.exists(), "CSV artifact should be written")
        assert_true(result.report_path.exists(), "Report artifact should be written")

    print("analysis_mode_tests=passed")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
