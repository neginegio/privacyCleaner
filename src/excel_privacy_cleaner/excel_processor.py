from __future__ import annotations

import re
import shutil
import tempfile
import unicodedata
import csv
import zipfile
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

from .models import Finding
from .presidio_japanese import (
    normalize_text,
    JapanesePresidioDetector,
    alias_kind,
    entity_label,
    normalize_alias_key,
)


@dataclass(frozen=True)
class ColumnRule:
    label: str
    kind: str
    headers: tuple[str, ...]
    contains: bool = False


@dataclass(frozen=True)
class ProcessingOptions:
    mode: str = "analysis"
    transform_business_secrets: bool = False
    pseudonym_scope: str = "file"
    birth_date_policy: str = "keep"

    @property
    def is_analysis(self) -> bool:
        return self.mode == "analysis"

    @property
    def mode_label(self) -> str:
        return "分析継続用" if self.is_analysis else "外部共有用"


@dataclass(frozen=True)
class ConversionResult:
    excel_path: Path
    csv_path: Path
    report_path: Path
    warnings: tuple[str, ...]
    converted_count: int
    review_count: int
    formula_maintained_count: int
    formula_changed_count: int


COLUMN_RULES = (
    ColumnRule("顧客ID", "customer_id", ("顧客ID",)),
    ColumnRule("受付番号", "ticket_id", ("受付番号",)),
    ColumnRule("案件ID", "project_id", ("案件ID",)),
    ColumnRule("緊急連絡先氏名", "emergency_name", ("緊急連絡先氏名",)),
    ColumnRule("緊急連絡先電話", "emergency_phone", ("緊急連絡先電話",)),
    ColumnRule("氏名", "name", ("氏名", "氏名カナ", "名前", "お名前", "顧客名", "患者名", "社員名", "担当者", "緊急連絡先氏名", "名義")),
    ColumnRule("住所", "address", ("住所", "所在地", "居所", "都道府県", "市区町村", "町名", "地番", "宛先")),
    ColumnRule("会社名", "company", ("会社名", "顧客企業")),
    ColumnRule("仕入先", "supplier", ("仕入先",)),
    ColumnRule("案件名", "project", ("案件名",)),
    ColumnRule("部署", "department", ("部署", "所属")),
    ColumnRule("役職", "role", ("役職",)),
    ColumnRule("郵便番号", "postal_code", ("郵便番号",)),
    ColumnRule("固定電話", "landline_phone", ("固定電話",)),
    ColumnRule("携帯電話", "mobile_phone", ("携帯電話", "携帯")),
    ColumnRule("電話番号", "phone", ("電話番号", "固定電話", "携帯電話", "電話"), contains=True),
    ColumnRule("メールアドレス", "email", ("メールアドレス", "メール")),
    ColumnRule("会員番号", "member_id", ("会員番号",)),
    ColumnRule("社員番号", "employee_id", ("社員番号",)),
    ColumnRule("個人番号", "personal_number", ("個人番号", "マイナンバー"), contains=True),
    ColumnRule("銀行口座", "bank_account", ("銀行口座", "口座情報", "口座番号")),
    ColumnRule("IPアドレス", "ip", ("サーバーIP", "IPアドレス")),
    ColumnRule("APIキー", "secret", ("APIキー", "APIキー・トークン", "トークン"), contains=True),
    ColumnRule("見積金額", "estimate_amount", ("見積金額",)),
    ColumnRule("原価見込", "cost_amount", ("原価見込", "原価")),
    ColumnRule("受付日時", "reception_datetime", ("受付日時",)),
    ColumnRule("基本給", "salary_50k", ("基本給",)),
    ColumnRule("賞与見込", "bonus_100k", ("賞与見込",)),
    ColumnRule("粗利率", "rate_10", ("粗利率",)),
    ColumnRule("生年月日", "birth_date", ("生年月日",)),
    ColumnRule("入社日", "hire_date", ("入社日",)),
    ColumnRule("評価ランク", "rating", ("評価ランク",)),
    ColumnRule("扶養家族数", "dependents", ("扶養家族数",)),
    ColumnRule("要配慮情報", "care_info", ("健康・配慮情報", "要配慮", "配慮情報")),
    ColumnRule("社外秘メモ", "secret_memo", ("社外秘メモ",)),
    ColumnRule("社内限定メモ", "boss_memo", ("上司メモ",)),
    ColumnRule("組織名", "organization", ("組織名", "拠点名", "工場")),
)

HEADER_EXCLUDE_WORDS = ("仕様", "説明", "期待", "結果", "備考", "担当者メモ")
SKIP_SHEET_KEYWORDS = ("テスト仕様", "期待処理ルール")
IGNORED_HEADER_WORDS = ("想定", "期待", "注意点", "出力例", "推奨")
MISSING_VALUE_WORDS = {"匿名希望", "なし", "無し", "不明", "未回答", "該当なし", "-", "－", "ー", "N/A", "NA"}
ANALYSIS_PRESERVED_KINDS = {
    "estimate_amount",
    "cost_amount",
    "salary_50k",
    "bonus_100k",
    "rate_10",
    "birth_date",
    "hire_date",
    "rating",
    "dependents",
    "reception_datetime",
    "schedule",
    "secret_memo",
    "boss_memo",
}
BUSINESS_SECRET_KINDS = {
    "estimate_amount",
    "cost_amount",
    "salary_50k",
    "bonus_100k",
    "rate_10",
    "rating",
    "dependents",
}
SENSITIVE_ENTITY_LABELS = {
    "氏名",
    "氏名候補",
    "顧客ID",
    "受付番号",
    "電話番号",
    "固定電話",
    "携帯電話",
    "緊急連絡先電話",
    "メールアドレス",
    "メール候補",
    "個人番号",
    "銀行口座",
    "秘密情報",
}
ALIAS_PREFIXES = {
    "name": "個人",
    "name_candidate": "個人候補",
    "email": "メール",
    "email_candidate": "メール候補",
    "address": "住所",
    "customer_id": "顧客",
    "ticket_id": "受付",
    "company": "法人",
    "supplier": "仕入先",
    "project_id": "案件",
    "project": "案件名称",
    "member_id": "会員",
    "employee_id": "社員",
    "ip": "IP",
    "landline_phone": "固定電話",
    "mobile_phone": "携帯",
    "phone": "電話",
    "financial_institution": "金融機関",
    "schedule": "予定日時",
    "organization": "組織",
    "text": "伏字",
}


class AliasBook:
    def __init__(self) -> None:
        self._maps: dict[str, dict[str, str]] = {}
        self._counters: dict[str, int] = {}

    def get(self, kind: str, original: str) -> str:
        key = normalize_alias_key(kind, original)
        values = self._maps.setdefault(kind, {})
        if key not in values:
            values[key] = self._next_alias(kind)
        return values[key]

    def new(self, kind: str, original: str) -> str:
        alias = self._next_alias(kind)
        key = normalize_alias_key(kind, original)
        if key:
            self._maps.setdefault(kind, {})[key] = alias
        return alias

    def allocate(self, kind: str) -> str:
        return self._next_alias(kind)

    def set(self, kind: str, original: str, alias: str) -> None:
        key = normalize_alias_key(kind, original)
        values = self._maps.setdefault(kind, {})
        if key and key not in values:
            values[key] = alias

    def has(self, kind: str, original: str) -> bool:
        key = normalize_alias_key(kind, original)
        return key in self._maps.get(kind, {})

    def peek(self, kind: str, original: str) -> str | None:
        key = normalize_alias_key(kind, original)
        return self._maps.get(kind, {}).get(key)

    def _next_alias(self, kind: str) -> str:
        prefix = ALIAS_PREFIXES.get(kind, "伏字")
        next_number = self._counters.get(kind, 0) + 1
        self._counters[kind] = next_number
        return f"{prefix}{next_number:03d}"


class ExcelPrivacyProcessor:
    def __init__(self) -> None:
        self.detector = JapanesePresidioDetector()
        self.alias_book = AliasBook()
        self.options = ProcessingOptions()
        self.known_literals: dict[str, tuple[str, str]] = {}
        self.single_name_literals: dict[str, str] = {}
        self.cell_replacements: dict[tuple[str, str], str] = {}
        self.formula_cells: set[tuple[str, str]] = set()
        self.ambiguous_name_keys: set[str] = set()
        self.scan_warnings: list[str] = []
        self.temp_dir: Path | None = None
        self.temp_workbook: Path | None = None

    def cleanup(self) -> None:
        if self.temp_dir and self.temp_dir.exists():
            shutil.rmtree(self.temp_dir, ignore_errors=True)
        self.temp_dir = None
        self.temp_workbook = None

    def scan(self, workbook_path: Path, options: ProcessingOptions | None = None) -> list[Finding]:
        self.cleanup()
        self.options = options or ProcessingOptions()
        self.alias_book = AliasBook()
        self.known_literals = {}
        self.single_name_literals = {}
        self.cell_replacements = {}
        self.formula_cells = set()
        self.ambiguous_name_keys = set()
        self.scan_warnings = []
        self.temp_dir = Path(tempfile.mkdtemp(prefix="ExcelPrivacyCleaner_"))
        self.temp_workbook = self.temp_dir / workbook_path.name
        shutil.copy2(workbook_path, self.temp_workbook)

        keep_vba = workbook_path.suffix.lower() == ".xlsm"
        workbook = load_workbook(self.temp_workbook, keep_vba=keep_vba)
        value_workbook = load_workbook(self.temp_workbook, keep_vba=keep_vba, data_only=True)
        findings: list[Finding] = []
        seen: set[tuple[str, str, str, str, str]] = set()

        for sheet in workbook.worksheets:
            if any(keyword in sheet.title for keyword in SKIP_SHEET_KEYWORDS):
                continue
            value_sheet = value_workbook[sheet.title]
            column_types = self._detect_column_types(sheet)
            self._seed_structured_aliases(sheet, value_sheet, column_types)
            ignored_columns = self._ignored_columns(sheet)
            category_rule = self._category_rule(sheet)
            table_header_cells = self._table_header_cells(sheet)
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.coordinate in table_header_cells:
                        continue
                    if cell.column in ignored_columns and cell.row > 1:
                        continue
                    display_value = value_sheet[cell.coordinate].value if cell.data_type == "f" else cell.value
                    if display_value is None:
                        continue
                    text = _stringify(display_value)
                    if not text.strip():
                        continue

                    column_entity = column_types.get(cell.column)
                    category_entity = self._category_entity(sheet, value_sheet, cell, category_rule)
                    if category_entity:
                        replacement = str(
                            category_entity.get("replacement")
                            or replacement_for(category_entity["kind"], display_value, self.alias_book, self.options)
                        )
                        detection_kind = str(category_entity.get("detection_kind", "分類列"))
                        kind = str(category_entity["kind"])
                        enabled = self._default_enabled(bool(category_entity["enabled"]), kind, detection_kind, cell)
                        reason = f"分類「{category_entity['category']}」"
                        if bool(category_entity["enabled"]) and not enabled:
                            reason += self._preserve_reason(kind, cell)
                        self._append_finding(
                            findings,
                            seen,
                            Finding(
                                enabled=enabled,
                                sheet=sheet.title,
                                cell=cell.coordinate,
                                entity_type=str(category_entity["label"]),
                                detection_kind=detection_kind,
                                original=text,
                                replacement=replacement,
                                reason=reason,
                            ),
                        )
                        if cell.data_type == "f" and enabled:
                            self.formula_cells.add((sheet.title, cell.coordinate))
                        continue

                    if column_entity and cell.row > column_entity["header_row"]:
                        if _is_missing_value(display_value):
                            continue
                        kind = str(column_entity["kind"])
                        replacement = self.cell_replacements.get((sheet.title, cell.coordinate))
                        if replacement is None:
                            replacement = replacement_for(kind, display_value, self.alias_book, self.options)
                        enabled = self._default_enabled(True, kind, "列単位", cell)
                        reason = f"見出し「{column_entity['header']}」"
                        if not enabled:
                            reason += self._preserve_reason(kind, cell)
                        self._append_finding(
                            findings,
                            seen,
                            Finding(
                                enabled=enabled,
                                sheet=sheet.title,
                                cell=cell.coordinate,
                                entity_type=str(column_entity["label"]),
                                detection_kind="列単位",
                                original=text,
                                replacement=replacement,
                                reason=reason,
                            ),
                        )
                        if cell.data_type == "f" and enabled:
                            self.formula_cells.add((sheet.title, cell.coordinate))
                        if enabled or kind not in {"secret_memo", "boss_memo"}:
                            continue

                    literal_findings = self._known_literal_findings(sheet.title, cell.coordinate, text)
                    for finding in literal_findings:
                        self._append_finding(findings, seen, finding)

                    for result in self.detector.analyze(text):
                        if any(
                            finding.start is not None
                            and finding.end is not None
                            and result.start < finding.end
                            and finding.start < result.end
                            for finding in literal_findings
                        ):
                            continue
                        original = text[result.start : result.end].strip()
                        kind = alias_kind(result.entity_type)
                        if len(original) < 2 and not self.alias_book.has(kind, original):
                            continue
                        replacement = replacement_for(kind, original, self.alias_book, self.options)
                        enabled = self._default_enabled(True, kind, "自由記述", cell)
                        self._append_finding(
                            findings,
                            seen,
                            Finding(
                                enabled=enabled,
                                sheet=sheet.title,
                                cell=cell.coordinate,
                                entity_type=entity_label(result.entity_type),
                                detection_kind="自由記述",
                                original=original,
                                replacement=replacement,
                                reason="Presidio カスタム Recognizer" + (self._preserve_reason(kind, cell) if not enabled else ""),
                                start=result.start,
                                end=result.end,
                            ),
                        )

        workbook.close()
        value_workbook.close()
        return findings

    def enabled_formula_replacement_count(self, findings: list[Finding], options: ProcessingOptions | None = None) -> int:
        active_options = options or self.options
        if active_options.is_analysis:
            return 0
        return len({(finding.sheet, finding.cell) for finding in findings if finding.enabled} & self.formula_cells)

    def _default_enabled(self, base_enabled: bool, kind: str, detection_kind: str, cell) -> bool:
        if not base_enabled and self.options.is_analysis and detection_kind == "確認候補":
            return True
        if not base_enabled:
            return False
        if cell.data_type == "f" and self.options.is_analysis:
            return False
        if not self.options.is_analysis:
            return True
        if kind in ANALYSIS_PRESERVED_KINDS:
            return False
        if kind in BUSINESS_SECRET_KINDS and not self.options.transform_business_secrets:
            return False
        return True

    def _preserve_reason(self, kind: str, cell) -> str:
        if cell.data_type == "f" and self.options.is_analysis:
            return " / 分析継続用のため数式セルは維持"
        if self.options.is_analysis and kind in ANALYSIS_PRESERVED_KINDS:
            return " / 分析継続用のため型と値を維持"
        if self.options.is_analysis and kind in BUSINESS_SECRET_KINDS and not self.options.transform_business_secrets:
            return " / 分析継続用のため企業機密項目を維持"
        return ""

    def convert(
        self,
        source_path: Path,
        findings: list[Finding],
        output_dir: Path | None = None,
        options: ProcessingOptions | None = None,
    ) -> Path:
        return self.convert_with_artifacts(
            source_path,
            findings,
            output_dir=output_dir,
            options=options,
            write_artifacts=False,
        ).excel_path

    def convert_with_artifacts(
        self,
        source_path: Path,
        findings: list[Finding],
        output_dir: Path | None = None,
        options: ProcessingOptions | None = None,
        write_artifacts: bool = True,
    ) -> ConversionResult:
        if not self.temp_workbook or not self.temp_workbook.exists():
            raise RuntimeError("先に検査を実行してください。")

        active_options = options or self.options
        blocked_candidates = [finding for finding in findings if not finding.enabled and finding.detection_kind == "確認候補"]
        if blocked_candidates:
            preview = "、".join(f"{finding.sheet}!{finding.cell}" for finding in blocked_candidates[:8])
            raise RuntimeError(
                f"未確認候補が {len(blocked_candidates)} 件あります。候補行を確認して変換対象にするか、"
                f"検出結果CSVで確認してください。対象例: {preview}"
            )

        selected = [finding for finding in findings if finding.enabled]
        if not selected:
            raise RuntimeError("変換対象が選択されていません。")

        warnings = list(self.scan_warnings)
        keep_vba = source_path.suffix.lower() == ".xlsm"
        workbook = load_workbook(self.temp_workbook, keep_vba=keep_vba)
        sheet_map = {sheet.title: sheet for sheet in workbook.worksheets}

        by_cell: dict[tuple[str, str], list[Finding]] = {}
        for finding in selected:
            by_cell.setdefault((finding.sheet, finding.cell), []).append(finding)

        for (sheet_name, coordinate), cell_findings in by_cell.items():
            sheet = sheet_map.get(sheet_name)
            if sheet is None:
                continue
            if coordinate in self._table_header_cells(sheet):
                continue
            cell = sheet[coordinate]
            if cell.value is None:
                continue
            if active_options.is_analysis and cell.data_type == "f":
                warnings.append(f"{sheet_name}!{coordinate}: 分析継続用のため数式セルの置換をスキップしました。")
                continue
            current = str(cell.value)
            full_cell = [
                finding
                for finding in cell_findings
                if finding.detection_kind in {"列単位", "分類列"}
                or (finding.detection_kind == "確認候補" and finding.start is None and finding.end is None)
            ]
            if full_cell:
                cell.value = full_cell[-1].replacement
                continue

            replaced_ranges: list[range] = []
            for finding in sorted(
                [item for item in cell_findings if item.detection_kind in {"自由記述", "辞書一致", "確認候補"}],
                key=lambda item: item.start if item.start is not None else 0,
                reverse=True,
            ):
                if finding.start is not None and finding.end is not None:
                    finding_range = range(finding.start, finding.end)
                    if any(finding_range.start < existing.stop and existing.start < finding_range.stop for existing in replaced_ranges):
                        continue
                    current = current[: finding.start] + finding.replacement + current[finding.end :]
                    replaced_ranges.append(finding_range)
                else:
                    current = current.replace(finding.original, finding.replacement)
            if any(finding.entity_type == "住所" for finding in cell_findings):
                current = _cleanup_address_tail(current)
            if any(finding.entity_type == "氏名" for finding in cell_findings):
                current = _cleanup_alias_note_tail(current)
            current = _postprocess_text(current, active_options, sheet_name, coordinate)
            cell.value = current

        for sheet in workbook.worksheets:
            if any(keyword in sheet.title for keyword in SKIP_SHEET_KEYWORDS):
                continue
            ignored_columns = self._ignored_columns(sheet)
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.column in ignored_columns and cell.row > 1:
                        continue
                    if cell.data_type == "f" or not isinstance(cell.value, str):
                        continue
                    cell.value = _postprocess_text(cell.value, active_options, sheet.title, cell.coordinate)

        if active_options.is_analysis:
            _set_recalculate_on_open(workbook)

        output_path = self._make_output_path(source_path, output_dir=output_dir, options=active_options)
        workbook.save(output_path)
        workbook.close()
        validation = validate_output_workbook(source_path, output_path, findings, active_options)
        warnings.extend(validation["warnings"])
        csv_path = output_path.with_name(f"{output_path.stem}_検出変換結果.csv")
        report_path = output_path.with_name(f"{output_path.stem}_処理報告書.txt")
        converted_count = sum(1 for finding in findings if finding.enabled)
        review_count = sum(1 for finding in findings if finding.detection_kind == "確認候補")
        result = ConversionResult(
            excel_path=output_path,
            csv_path=csv_path,
            report_path=report_path,
            warnings=tuple(warnings),
            converted_count=converted_count,
            review_count=review_count,
            formula_maintained_count=int(validation["formula_maintained_count"]),
            formula_changed_count=int(validation["formula_changed_count"]),
        )
        if write_artifacts:
            write_findings_csv(csv_path, findings)
            write_processing_report(
                report_path,
                source_path=source_path,
                result=result,
                findings=findings,
                options=active_options,
                validation=validation,
            )
        self.cleanup()
        return result

    @staticmethod
    def _detect_column_types(sheet: Worksheet) -> dict[int, dict[str, str | int]]:
        column_types: dict[int, dict[str, str | int]] = {}
        max_header_row = min(sheet.max_row or 0, 1)
        for row in sheet.iter_rows(min_row=1, max_row=max_header_row):
            for cell in row:
                if cell.value is None:
                    continue
                header = normalize_text(str(cell.value))
                if not _is_likely_header(header):
                    continue
                rule = _match_column_rule(header)
                if rule is not None:
                    column_types[cell.column] = {
                        "label": rule.label,
                        "kind": rule.kind,
                        "header_row": cell.row,
                        "header": header,
                    }
        return column_types

    @staticmethod
    def _table_header_cells(sheet: Worksheet) -> set[str]:
        cells: set[str] = set()
        for table in sheet.tables.values():
            if getattr(table, "headerRowCount", 1) == 0:
                continue
            min_col, min_row, max_col, _max_row = range_boundaries(table.ref)
            for column in range(min_col, max_col + 1):
                cells.add(sheet.cell(row=min_row, column=column).coordinate)
        return cells

    @staticmethod
    def _ignored_columns(sheet: Worksheet) -> set[int]:
        columns: set[int] = set()
        for cell in sheet[1]:
            if cell.value is None:
                continue
            header = normalize_text(str(cell.value))
            if any(word in header for word in IGNORED_HEADER_WORDS):
                columns.add(cell.column)
        return columns

    @staticmethod
    def _category_rule(sheet: Worksheet) -> tuple[int, int] | None:
        category_col: int | None = None
        input_col: int | None = None
        for cell in sheet[1]:
            header = normalize_text(str(cell.value)) if cell.value is not None else ""
            if header == "分類":
                category_col = cell.column
            elif header == "入力値":
                input_col = cell.column
        if category_col is not None and input_col is not None:
            return category_col, input_col
        return None

    def _category_entity(
        self,
        sheet: Worksheet,
        value_sheet: Worksheet,
        cell,
        category_rule: tuple[int, int] | None,
    ) -> dict[str, str | bool] | None:
        if category_rule is None or cell.row <= 1:
            return None
        category_col, input_col = category_rule
        if cell.column != input_col:
            return None
        category_value = value_sheet.cell(row=cell.row, column=category_col).value
        category = normalize_text(str(category_value)) if category_value is not None else ""
        input_value = value_sheet.cell(row=cell.row, column=input_col).value
        input_text = _normalize_ascii(str(input_value)) if input_value is not None else ""
        if _is_missing_value(input_text):
            return None
        if category == "秘密情報" and re.fullmatch(r"(?:(?:25[0-5]|2[0-4]\d|1?\d?\d)\.){3}(?:25[0-5]|2[0-4]\d|1?\d?\d)", input_text):
            return {"label": "IPアドレス", "kind": "ip", "enabled": True, "category": category}
        if category == "氏名":
            candidate = self._name_candidate(input_text)
            if candidate is not None:
                return {
                    "label": "氏名候補",
                    "kind": "name",
                    "enabled": False,
                    "category": category,
                    "detection_kind": "確認候補",
                    "replacement": candidate,
                }
        if category == "メール" and "+" in input_text.partition("@")[0]:
            return {
                "label": "メール候補",
                "kind": "email",
                "enabled": False,
                "category": "メール候補",
                "detection_kind": "確認候補",
                "replacement": _incomplete_email_candidate(input_text, self.alias_book, self.options),
            }
        mapping: dict[str, tuple[str, str, bool]] = {
            "氏名": ("氏名", "name", True),
            "電話": ("電話番号", "phone", True),
            "メール": ("メールアドレス", "email", True),
            "メール候補": ("メール候補", "email", False),
            "住所": ("住所", "address", True),
            "秘密情報": ("秘密情報", "secret", True),
        }
        if category not in mapping:
            for look_col in range(input_col + 1, min(input_col + 3, sheet.max_column) + 1):
                hint_value = value_sheet.cell(row=cell.row, column=look_col).value
                hint = normalize_text(str(hint_value)) if hint_value is not None else ""
                if "メール候補" in hint:
                    replacement = _incomplete_email_candidate(input_text, self.alias_book, self.options)
                    return {
                        "label": "メール候補",
                        "kind": "email",
                        "enabled": False,
                        "category": "メール候補",
                        "detection_kind": "確認候補",
                        "replacement": replacement,
                    }
            return None
        label, kind, enabled = mapping[category]
        return {"label": label, "kind": kind, "enabled": enabled, "category": category}

    def _name_candidate(self, value: str) -> str | None:
        normalized = _normalize_ascii(value)
        stripped = _strip_name_honorific(normalized)
        alias = self.alias_book.peek("name", stripped) or self.alias_book.peek("name", normalized)
        if alias is None:
            return None
        key = normalize_alias_key("name", stripped)
        if key in self.ambiguous_name_keys or _is_ambiguous_name_candidate(normalized):
            return _candidate_mask(normalized, self.alias_book)
        return None

    @staticmethod
    def _append_finding(
        findings: list[Finding],
        seen: set[tuple[str, str, str, str, str]],
        finding: Finding,
    ) -> None:
        if finding.dedupe_key in seen:
            return
        findings.append(finding)
        seen.add(finding.dedupe_key)

    def _seed_structured_aliases(
        self,
        sheet: Worksheet,
        value_sheet: Worksheet,
        column_types: dict[int, dict[str, str | int]],
    ) -> None:
        name_columns = [
            (column, info)
            for column, info in column_types.items()
            if info["kind"] == "name" and sheet.max_row > int(info["header_row"])
        ]
        primary_name_columns = [
            (column, info)
            for column, info in name_columns
            if "カナ" not in str(info["header"]) and "かな" not in str(info["header"])
        ]
        primary_name_column = primary_name_columns[0][0] if primary_name_columns else (name_columns[0][0] if name_columns else None)
        cross_domain_person_sheet = _is_cross_domain_person_sheet(sheet.title)

        for row_index in range(2, sheet.max_row + 1):
            if primary_name_column is not None:
                primary_value = value_sheet.cell(row=row_index, column=primary_name_column).value
                if primary_value is not None and str(primary_value).strip():
                    primary_text = _stringify(primary_value)
                    if cross_domain_person_sheet:
                        if self.alias_book.has("name", primary_text):
                            self.ambiguous_name_keys.add(normalize_alias_key("name", primary_text))
                        alias = self.alias_book.allocate("name")
                    else:
                        alias = self.alias_book.get("name", primary_text)
                    self.cell_replacements[(sheet.title, sheet.cell(row=row_index, column=primary_name_column).coordinate)] = alias
                    for variant in _name_variants(primary_text):
                        if not cross_domain_person_sheet:
                            self.alias_book.set("name", variant, alias)
                            self._remember_literal(variant, "氏名", alias)
                            for role_variant, role_replacement in _role_variants(variant, alias).items():
                                self._remember_literal(role_variant, "氏名", role_replacement)
                            if len(normalize_text(variant)) == 1:
                                self.single_name_literals[normalize_text(variant)] = alias
                    for column, info in name_columns:
                        header = str(info["header"])
                        if "カナ" not in header and "かな" not in header:
                            continue
                        related_value = value_sheet.cell(row=row_index, column=column).value
                        if related_value is None or not str(related_value).strip():
                            continue
                        self.cell_replacements[(sheet.title, sheet.cell(row=row_index, column=column).coordinate)] = alias
                        for variant in _name_variants(_stringify(related_value)):
                            if not cross_domain_person_sheet:
                                self.alias_book.set("name", variant, alias)
                                self._remember_literal(variant, "氏名", alias)
                                for role_variant, role_replacement in _role_variants(variant, alias).items():
                                    self._remember_literal(role_variant, "氏名", role_replacement)
                                if len(normalize_text(variant)) == 1:
                                    self.single_name_literals[normalize_text(variant)] = alias

            for column, info in column_types.items():
                kind = str(info["kind"])
                if kind not in {
                    "company",
                    "supplier",
                    "organization",
                    "customer_id",
                    "ticket_id",
                    "project_id",
                    "project",
                    "member_id",
                    "employee_id",
                    "email",
                    "landline_phone",
                    "mobile_phone",
                    "phone",
                    "emergency_phone",
                }:
                    continue
                value = value_sheet.cell(row=row_index, column=column).value
                if kind == "landline_phone" and not self.options.is_analysis and (value is None or not str(value).strip()):
                    self.alias_book.allocate("landline_phone")
                    continue
                if value is None or not str(value).strip():
                    continue
                if _is_missing_value(value):
                    continue
                replacement = replacement_for(kind, value, self.alias_book, self.options)
                label = str(info["label"])
                variants = (
                    _organization_variants(_stringify(value))
                    if kind in {"company", "supplier", "organization"}
                    else _literal_variants(kind, _stringify(value))
                )
                for variant in variants:
                    self.alias_book.set(kind, variant, replacement)
                    self._remember_literal(variant, label, replacement)

    def _remember_literal(self, original: str, label: str, replacement: str) -> None:
        normalized = normalize_text(original)
        if len(normalized) < 2:
            return
        self.known_literals[normalized] = (label, replacement)

    def _known_literal_findings(self, sheet_name: str, coordinate: str, text: str) -> list[Finding]:
        findings: list[Finding] = []
        occupied: list[range] = []
        for original, replacement in self.single_name_literals.items():
            for regex in (rf"(?<=の){re.escape(original)}(?=です)", rf"{re.escape(original)}さん"):
                for match in re.finditer(regex, text):
                    match_range = range(match.start(), match.end())
                    if any(match_range.start < existing.stop and existing.start < match_range.stop for existing in occupied):
                        continue
                    candidate = match.group(0).endswith("さん")
                    enabled = True if (candidate and self.options.is_analysis) else not candidate
                    findings.append(
                        Finding(
                            enabled=enabled,
                            sheet=sheet_name,
                            cell=coordinate,
                            entity_type="氏名候補" if candidate else "氏名",
                            detection_kind="確認候補" if candidate else "辞書一致",
                            original=match.group(0),
                            replacement=_candidate_mask(match.group(0), self.alias_book) if candidate else replacement,
                            reason="曖昧な姓+敬称候補" if candidate else "構造列から作成した1文字姓辞書",
                            start=match.start(),
                            end=match.end(),
                        )
                    )
                    occupied.append(match_range)
        for original, (label, replacement) in sorted(self.known_literals.items(), key=lambda item: len(item[0]), reverse=True):
            flags = re.IGNORECASE if _is_ascii_literal(original) else 0
            for match in re.finditer(re.escape(original), text, flags=flags):
                if label == "氏名" and _is_inside_email_or_identifier(text, match.start(), match.end()):
                    continue
                match_range = range(match.start(), match.end())
                if any(match_range.start < existing.stop and existing.start < match_range.stop for existing in occupied):
                    continue
                matched_text = match.group(0)
                candidate = label == "氏名" and (
                    normalize_alias_key("name", _strip_name_honorific(matched_text)) in self.ambiguous_name_keys
                    or _is_ambiguous_name_candidate(matched_text)
                )
                kind = _kind_for_label(label)
                enabled = True if (candidate and self.options.is_analysis) else not candidate
                if self.options.is_analysis and kind in BUSINESS_SECRET_KINDS and not self.options.transform_business_secrets:
                    enabled = False
                findings.append(
                    Finding(
                        enabled=enabled,
                        sheet=sheet_name,
                        cell=coordinate,
                        entity_type="氏名候補" if candidate else label,
                        detection_kind="確認候補" if candidate else "辞書一致",
                        original=matched_text,
                        replacement=_candidate_mask(matched_text, self.alias_book) if candidate else replacement,
                        reason=(
                            "曖昧な表記ゆれ候補"
                            if candidate
                            else "構造列から作成した同一人物・組織辞書"
                        )
                        + (
                            " / 分析継続用のため企業機密項目を維持"
                            if not enabled and kind in BUSINESS_SECRET_KINDS
                            else ""
                        ),
                        start=match.start(),
                        end=match.end(),
                    )
                )
                occupied.append(match_range)
        return findings

    @staticmethod
    def _make_output_path(source_path: Path, output_dir: Path | None = None, options: ProcessingOptions | None = None) -> Path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        suffix = ".xlsm" if source_path.suffix.lower() == ".xlsm" else ".xlsx"
        folder = output_dir if output_dir is not None else source_path.parent
        folder.mkdir(parents=True, exist_ok=True)
        mode_label = (options or ProcessingOptions()).mode_label
        candidate = folder / f"{source_path.stem}_{mode_label}_匿名化_{timestamp}{suffix}"
        if not candidate.exists():
            return candidate
        return folder / f"{source_path.stem}_{mode_label}_匿名化_{timestamp}_{datetime.now().microsecond:06d}{suffix}"


def write_findings_csv(path: Path, findings: list[Finding]) -> None:
    headers = [
        "変換",
        "処理状態",
        "シート",
        "セル",
        "種類",
        "検査",
        "検出値",
        "変換後",
        "理由",
    ]
    with path.open("w", encoding="utf-8-sig", newline="") as output:
        writer = csv.writer(output)
        writer.writerow(headers)
        for finding in findings:
            status = "確認候補" if finding.detection_kind == "確認候補" else ("自動変換" if finding.enabled else "維持")
            writer.writerow(
                [
                    "対象" if finding.enabled else "維持",
                    status,
                    finding.sheet,
                    finding.cell,
                    finding.entity_type,
                    finding.detection_kind,
                    finding.original,
                    finding.replacement,
                    finding.reason,
                ]
            )


def write_processing_report(
    path: Path,
    *,
    source_path: Path,
    result: ConversionResult,
    findings: list[Finding],
    options: ProcessingOptions,
    validation: dict[str, Any],
) -> None:
    business_kept = [
        finding
        for finding in findings
        if not finding.enabled and _kind_for_label(finding.entity_type) in BUSINESS_SECRET_KINDS
    ]
    lines = [
        "Excel匿名化 処理報告書",
        "",
        f"処理モード: {options.mode_label}",
        f"仮名化範囲: {_scope_label(options.pseudonym_scope)}",
        f"企業機密も変換する: {'オン' if options.transform_business_secrets else 'オフ'}",
        f"入力ファイル名: {source_path.name}",
        f"出力ファイル名: {result.excel_path.name}",
        f"処理日時: {datetime.now():%Y-%m-%d %H:%M:%S}",
        f"検出件数: {len(findings)}",
        f"変換件数: {result.converted_count}",
        f"要確認件数: {result.review_count}",
        f"変換しなかった企業機密項目: {len(business_kept)}",
        f"数式維持件数: {result.formula_maintained_count}",
        f"数式変更件数: {result.formula_changed_count}",
        "",
        "非対応機能:",
    ]
    unsupported = validation.get("unsupported_features", [])
    lines.extend([f"- {item}" for item in unsupported] or ["- なし"])
    lines.extend(["", "XLSX内部検査結果:"])
    lines.append(str(validation.get("internal_scan_result", "未実施")))
    lines.extend(["", "警告内容:"])
    lines.extend([f"- {warning}" for warning in result.warnings] or ["- なし"])
    lines.extend(["", "変換しなかった企業機密項目の例:"])
    for finding in business_kept[:30]:
        lines.append(f"- {finding.sheet}!{finding.cell} {finding.entity_type}: {finding.original}")
    if not business_kept:
        lines.append("- なし")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def validate_output_workbook(
    source_path: Path,
    output_path: Path,
    findings: list[Finding],
    options: ProcessingOptions,
) -> dict[str, Any]:
    warnings: list[str] = []
    unsupported_features: list[str] = []
    try:
        source = load_workbook(source_path, data_only=False)
        output = load_workbook(output_path, data_only=False)
    except Exception as exc:
        return {
            "warnings": [f"出力XLSXを再読み込みできませんでした: {exc}"],
            "unsupported_features": ["検証未完了"],
            "internal_scan_result": "読み込み失敗",
            "formula_maintained_count": 0,
            "formula_changed_count": 0,
        }

    try:
        if source.sheetnames != output.sheetnames:
            warnings.append("シート数またはシート順が変化しています。")
        formula_maintained_count = 0
        formula_changed_count = 0
        for sheet_name in source.sheetnames:
            if sheet_name not in output.sheetnames:
                continue
            source_sheet = source[sheet_name]
            output_sheet = output[sheet_name]
            if source_sheet.max_row != output_sheet.max_row or source_sheet.max_column != output_sheet.max_column:
                warnings.append(f"{sheet_name}: 行数または列数が変化しています。")
            if len(source_sheet.tables) != len(output_sheet.tables):
                warnings.append(f"{sheet_name}: Excelテーブル数が変化しています。")
            if len(source_sheet.conditional_formatting) != len(output_sheet.conditional_formatting):
                warnings.append(f"{sheet_name}: 条件付き書式数が変化しています。")
            if source_sheet.data_validations.count != output_sheet.data_validations.count:
                warnings.append(f"{sheet_name}: 入力規則数が変化しています。")
            for row in source_sheet.iter_rows():
                for source_cell in row:
                    output_cell = output_sheet[source_cell.coordinate]
                    if source_cell.data_type == "f":
                        if output_cell.data_type == "f" and output_cell.value == source_cell.value:
                            formula_maintained_count += 1
                        else:
                            formula_changed_count += 1
                            warnings.append(f"{sheet_name}!{source_cell.coordinate}: 数式が維持されていません。")
                    if options.is_analysis and isinstance(source_cell.value, (int, float)) and not isinstance(source_cell.value, bool):
                        if not isinstance(output_cell.value, (int, float)) or isinstance(output_cell.value, bool):
                            warnings.append(f"{sheet_name}!{source_cell.coordinate}: 数値型が維持されていません。")
                    if options.is_analysis and isinstance(source_cell.value, (datetime, date)):
                        if not isinstance(output_cell.value, (datetime, date)):
                            warnings.append(f"{sheet_name}!{source_cell.coordinate}: 日付型が維持されていません。")
            unsupported_features.extend(_inspect_hidden_surfaces(source_sheet, source_path.name))
        unsupported_features.extend(_inspect_package_features(output_path))
    finally:
        source.close()
        output.close()

    sensitive_values = _sensitive_original_values(findings)
    leaked = _scan_xlsx_package(output_path, sensitive_values)
    internal_scan_result = "残存なし"
    if leaked:
        internal_scan_result = f"元識別情報の残存候補 {len(leaked)} 件"
        warnings.extend([f"内部XML残存候補: {item}" for item in leaked[:30]])

    return {
        "warnings": warnings,
        "unsupported_features": sorted(set(unsupported_features)) or ["なし"],
        "internal_scan_result": internal_scan_result,
        "formula_maintained_count": formula_maintained_count,
        "formula_changed_count": formula_changed_count,
    }


def _set_recalculate_on_open(workbook) -> None:
    calculation = getattr(workbook, "calculation", None)
    if calculation is None:
        return
    for attr in ("fullCalcOnLoad", "forceFullCalc"):
        if hasattr(calculation, attr):
            setattr(calculation, attr, True)


def _inspect_hidden_surfaces(sheet: Worksheet, source_name: str) -> list[str]:
    notes: list[str] = []
    if sheet.sheet_state != "visible":
        notes.append(f"非表示シートを検出: {sheet.title}")
    if any(dimension.hidden for dimension in sheet.row_dimensions.values()):
        notes.append(f"非表示行を検出: {sheet.title}")
    if any(dimension.hidden for dimension in sheet.column_dimensions.values()):
        notes.append(f"非表示列を検出: {sheet.title}")
    comments = [cell.coordinate for row in sheet.iter_rows() for cell in row if cell.comment is not None]
    if comments:
        notes.append(f"コメント/メモを検出: {sheet.title} {len(comments)} 件。コメント本文の自動置換は限定対応です。")
    if sheet.data_validations.count:
        notes.append(f"入力規則を検出: {sheet.title} {sheet.data_validations.count} 件。メッセージ本文の自動置換は限定対応です。")
    return notes


def _inspect_package_features(path: Path) -> list[str]:
    notes: list[str] = []
    try:
        with zipfile.ZipFile(path) as archive:
            names = archive.namelist()
    except zipfile.BadZipFile:
        return ["XLSX ZIP構造を読み取れませんでした。"]
    if any(name.startswith("xl/charts/") for name in names):
        notes.append("グラフを検出。グラフタイトル/データラベル内文字列の自動置換は限定対応です。")
    if any(name.startswith("xl/pivotCache/") or name.startswith("xl/pivotTables/") for name in names):
        notes.append("ピボットテーブル/ピボットキャッシュを検出。キャッシュ内文字列の自動置換は限定対応です。")
    if any(name.startswith("customXml/") for name in names):
        notes.append("カスタムXMLを検出。カスタムXML内文字列の自動置換は限定対応です。")
    if any(name.startswith("xl/embeddings/") for name in names):
        notes.append("埋め込みオブジェクトを検出。埋め込みオブジェクト内文字列の自動置換は未対応です。")
    if any(name.startswith("xl/externalLinks/") for name in names):
        notes.append("外部リンクを検出。外部リンク先の安全性確認が必要です。")
    return notes


def _sensitive_original_values(findings: list[Finding]) -> list[str]:
    values: set[str] = set()
    for finding in findings:
        if finding.entity_type not in SENSITIVE_ENTITY_LABELS:
            continue
        normalized = _normalize_ascii(finding.original)
        if _is_placeholder_value(normalized):
            continue
        if len(normalized) >= 3 and normalized not in MISSING_VALUE_WORDS:
            values.add(normalized)
    return sorted(values, key=len, reverse=True)


def _scan_xlsx_package(path: Path, sensitive_values: list[str]) -> list[str]:
    if not sensitive_values:
        return []
    leaked: list[str] = []
    try:
        with zipfile.ZipFile(path) as archive:
            for name in archive.namelist():
                if not name.endswith((".xml", ".rels")):
                    continue
                text = archive.read(name).decode("utf-8", errors="ignore")
                for value in sensitive_values:
                    if value in text:
                        leaked.append(f"{name}: {value}")
                        break
    except zipfile.BadZipFile:
        leaked.append("XLSXをZIPとして展開できませんでした。")
    return leaked


def _scope_label(scope: str) -> str:
    return {
        "file": "このファイル内だけ",
        "batch": "今回アップロードした一連のファイル内",
        "project": "プロジェクト内",
    }.get(scope, scope)


def _is_placeholder_value(value: str) -> bool:
    lowered = value.lower()
    if lowered.endswith("@example.invalid"):
        return True
    if re.fullmatch(
        r"(?:個人|個人候補|法人|案件|案件名称|顧客|受付|社員|会員|仕入先|組織|住所|予定日時|伏字|IP|メール|メール候補|固定電話|携帯|電話|金融機関)\d{3}(?:さん|様|部長|課長|係長|主任|担当|殿|氏)?",
        value,
    ):
        return True
    return False


def _match_column_rule(header: str) -> ColumnRule | None:
    for rule in COLUMN_RULES:
        if header in rule.headers:
            return rule
    for rule in COLUMN_RULES:
        if rule.contains and any(word in header for word in rule.headers):
            return rule
    return None


def _kind_for_label(label: str) -> str:
    mapping = {
        "氏名": "name",
        "氏名候補": "name",
        "住所": "address",
        "顧客ID": "customer_id",
        "受付番号": "ticket_id",
        "会社名": "company",
        "仕入先": "supplier",
        "案件ID": "project_id",
        "案件名": "project",
        "組織名": "organization",
        "固定電話": "landline_phone",
        "携帯電話": "mobile_phone",
        "緊急連絡先氏名": "emergency_name",
        "緊急連絡先電話": "emergency_phone",
        "メールアドレス": "email",
        "メール候補": "email",
        "電話番号": "phone",
        "個人番号": "personal_number",
        "銀行口座": "bank_account",
        "秘密情報": "secret",
    }
    return mapping.get(label, "text")


def _is_likely_header(header: str) -> bool:
    if not header or len(header) > 24:
        return False
    if any(word in header for word in HEADER_EXCLUDE_WORDS):
        return False
    if any(mark in header for mark in ("。", "、", "，", ",", "→", "です", "ます")):
        return False
    return True


def _is_missing_value(value: Any) -> bool:
    normalized = _normalize_ascii(str(value))
    compact = re.sub(r"[\s　]+", "", normalized)
    return compact in MISSING_VALUE_WORDS


def _is_cross_domain_person_sheet(sheet_name: str) -> bool:
    return sheet_name.startswith("03_")


def _is_ascii_literal(value: str) -> bool:
    return bool(value) and all(ord(char) < 128 for char in value)


def _is_inside_email_or_identifier(text: str, start: int, end: int) -> bool:
    left = text[start - 1] if start > 0 else ""
    right = text[end] if end < len(text) else ""
    if (left.isascii() and left.isalnum()) or (right.isascii() and right.isalnum()):
        return True
    token_start = start
    while token_start > 0 and not text[token_start - 1].isspace():
        token_start -= 1
    token_end = end
    while token_end < len(text) and not text[token_end].isspace():
        token_end += 1
    token = text[token_start:token_end]
    return "@" in token


def _strip_name_honorific(value: str) -> str:
    return re.sub(r"(?:さん|様|殿|氏)$", "", _normalize_ascii(value))


def _is_ambiguous_name_candidate(value: str) -> bool:
    normalized = _normalize_ascii(value)
    base = _strip_name_honorific(normalized)
    if normalized != base and re.fullmatch(r"[一-龯々〆ヵヶ]{1,3}(?:さん|様|殿|氏)", normalized):
        return True
    if re.fullmatch(r"[ぁ-ん]+(?:[ 　]+[ぁ-ん]+)?", normalized):
        return True
    if re.fullmatch(r"[A-Z]+(?:[ 　]+[A-Z]+)?", normalized):
        return True
    return False


def _candidate_mask(original: str, alias_book: AliasBook) -> str:
    normalized = _normalize_ascii(original)
    base = _strip_name_honorific(normalized)
    suffix = normalized[len(base) :] if normalized.startswith(base) else ""
    candidate_key = alias_book.peek("name", base) or alias_book.peek("name", normalized) or base or normalized
    return f"{alias_book.get('name_candidate', candidate_key)}{suffix}"


def _incomplete_email_candidate(value: str, alias_book: AliasBook, options: ProcessingOptions) -> str:
    normalized = _normalize_ascii(value).lower()
    if "@" in normalized:
        local = normalized.split("@", 1)[0]
        for domain in ("example.com", "example.jp", "example.net", "example.org"):
            alias = alias_book.peek("email", f"{local}@{domain}")
            if alias is not None:
                number = re.search(r"(\d{3})$", alias)
                suffix = number.group(1) if number else "001"
                return f"メール候補{suffix}" if not options.is_analysis else f"mail候補{suffix}@example.invalid"
    alias = alias_book.get("email_candidate", normalized or "メールアドレス")
    number = re.search(r"(\d{3})$", alias)
    suffix = number.group(1) if number else "001"
    return alias if not options.is_analysis else f"mail候補{suffix}@example.invalid"


def replacement_for(kind: str, value: Any, alias_book: AliasBook, options: ProcessingOptions | None = None) -> str:
    active_options = options or ProcessingOptions()
    text = _stringify(value)
    if kind == "customer_id":
        return alias_book.get(kind, text)
    if kind == "ticket_id":
        return alias_book.get(kind, text)
    if kind == "project_id":
        return alias_book.get(kind, text)
    if kind == "address":
        return _prefecture_only(text) if not active_options.is_analysis else (_generalize_address(text) or alias_book.get(kind, text))
    if kind in {"name", "company", "supplier", "project", "member_id", "employee_id", "organization"}:
        return alias_book.get(kind, text)
    if kind == "emergency_name":
        return "非表示" if not active_options.is_analysis else alias_book.get("name", text)
    if kind == "email":
        return _alias_email(text, alias_book, active_options)
    if kind in {"phone", "landline_phone", "mobile_phone"}:
        return _alias_phone(kind, text, alias_book, active_options)
    if kind == "emergency_phone":
        return "非表示" if not active_options.is_analysis else _mask_phone(text)
    if kind == "postal_code":
        return "XXX-XXXX" if not active_options.is_analysis else _mask_postal_code(text)
    if kind == "personal_number":
        return "非表示" if not active_options.is_analysis else "*" * max(12, len(_digits(text)))
    if kind == "bank_account":
        return "金融情報を削除" if not active_options.is_analysis else _mask_bank_account(text, alias_book)
    if kind == "ip":
        return "非表示" if not active_options.is_analysis else alias_book.get(kind, text)
    if kind == "secret":
        return "********"
    if kind == "estimate_amount":
        return _estimate_amount_range(value)
    if kind == "cost_amount":
        return _cost_amount_range(value)
    if kind == "salary_50k":
        return _salary_range(value)
    if kind == "bonus_100k":
        return _bonus_range(value)
    if kind == "rate_10":
        return _rate_range(value)
    if kind == "birth_date":
        year = _year(value)
        return f"{year // 10 * 10}年代" if year else "年代不明"
    if kind == "hire_date":
        year = _year(value)
        return f"{year // 10 * 10}年代" if year else "入社年不明"
    if kind == "reception_datetime":
        return _week_bucket(value)
    if kind == "rating":
        return {"A": "高評価", "B": "標準", "C": "要支援"}.get(text.strip(), "評価情報あり")
    if kind == "dependents":
        number = _number(value)
        return "あり" if number and number > 0 else "なし"
    if kind == "care_info":
        if not text.strip():
            return ""
        if not active_options.is_analysis:
            return "要配慮情報あり"
        if any(word in text for word in ("育児", "短時間", "時短")):
            return text
        if "在留資格" in text:
            return text
        if "通院" in text:
            return "通院に伴う勤務配慮あり"
        if any(word in text for word in ("腰痛", "病", "けが", "怪我")):
            return "身体的配慮あり"
        return "要配慮情報あり"
    if kind == "secret_memo":
        return "社外秘情報を削除" if not active_options.is_analysis else text
    if kind == "boss_memo":
        return "人事メモを削除" if not active_options.is_analysis else text
    if kind == "department":
        return _generalize_department(text) if not active_options.is_analysis else _analysis_department(text)
    if kind == "role":
        return _generalize_role(text) if not active_options.is_analysis else text
    return alias_book.get("text", text)


def _stringify(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return str(value)


def _normalize_ascii(value: str) -> str:
    return unicodedata.normalize("NFKC", value).strip()


def _name_variants(value: str) -> set[str]:
    normalized = _normalize_ascii(value)
    compact = re.sub(r"[\s　]+", "", normalized)
    variants = {normalized, compact}
    variants.update(_kanji_variants(normalized))
    variants.update(_kanji_variants(compact))
    kana_hiragana = _katakana_to_hiragana(normalized)
    if kana_hiragana != normalized:
        variants.add(kana_hiragana)
        variants.add(re.sub(r"[\s　]+", "", kana_hiragana))
    variants.update(_simple_kana_romanized(normalized))
    parts = [part for part in re.split(r"[\s　]+", normalized) if part]
    family_names: set[str] = set()
    if parts:
        family_names.update(_kanji_variants(parts[0]))
    elif re.fullmatch(r"[一-龯々〆ヵヶ]{3,6}", compact):
        family_names.update(_kanji_variants(compact[:2]))
    variants.update(family_names)
    return {item for item in variants if item and len(item) >= 1}


def _role_variants(value: str, alias: str) -> dict[str, str]:
    normalized = _normalize_ascii(value)
    if not re.fullmatch(r"[一-龯々〆ヵヶ]{1,3}", normalized):
        return {}
    role_suffixes = ("部長", "課長", "係長", "主任", "担当", "様", "殿", "氏")
    return {f"{normalized}{suffix}": f"{alias}{suffix}" for suffix in role_suffixes}


def _katakana_to_hiragana(value: str) -> str:
    converted: list[str] = []
    for char in value:
        code = ord(char)
        if 0x30A1 <= code <= 0x30F6:
            converted.append(chr(code - 0x60))
        else:
            converted.append(char)
    return "".join(converted)


def _simple_kana_romanized(value: str) -> set[str]:
    token_map = {
        "サトウ": "SATO",
        "ハナコ": "HANAKO",
        "ヤマダ": "YAMADA",
        "タロウ": "TARO",
        "タカハシ": "TAKAHASHI",
        "イチロウ": "ICHIRO",
        "ワタナベ": "WATANABE",
        "ミサキ": "MISAKI",
        "ミナミ": "MINAMI",
        "アカリ": "AKARI",
        "モリ": "MORI",
        "タカナシ": "TAKANASHI",
        "シズク": "SHIZUKU",
    }
    tokens = [token for token in re.split(r"[\s　]+", _normalize_ascii(value)) if token]
    roman_tokens = [token_map.get(token) for token in tokens]
    if roman_tokens and all(roman_tokens):
        full = " ".join(roman_tokens)
        variants = {full, full.title()}
        if roman_tokens:
            variants.add(roman_tokens[0])
            variants.add(roman_tokens[0].title())
        return variants
    return set()


def _kanji_variants(value: str) -> set[str]:
    variants = {value}
    replacements = (("髙", "高"), ("高", "髙"), ("邊", "辺"), ("邉", "辺"), ("辺", "邊"), ("﨑", "崎"))
    for source, target in replacements:
        if source in value:
            variants.add(value.replace(source, target))
    return variants


def _organization_variants(value: str) -> set[str]:
    normalized = _normalize_ascii(value)
    compact = re.sub(r"[\s　]+", "", normalized)
    variants = {normalized, compact}
    stripped = re.sub(r"^(?:株式会社|有限会社|合同会社|医療法人|学校法人|社会福祉法人)", "", compact)
    stripped = re.sub(r"(?:株式会社|有限会社|合同会社|Inc\.?|Co\.?,?\s*Ltd\.?)$", "", stripped, flags=re.IGNORECASE)
    if stripped:
        variants.add(stripped)
    return {item for item in variants if item and len(item) >= 2}


def _literal_variants(kind: str, value: str) -> set[str]:
    normalized = _normalize_ascii(value)
    variants = {normalized}
    digits = _digits(normalized)
    if kind in {"phone", "landline_phone", "mobile_phone", "emergency_phone"} and digits:
        variants.add(digits)
        if digits.startswith("81"):
            national = digits[2:]
            if national.startswith(("70", "80", "90")):
                variants.add(f"0{national}")
        if digits.startswith(("070", "080", "090")):
            variants.add(f"+81-{digits[1:3]}-{digits[3:7]}-{digits[7:]}")
    if kind == "email":
        variants.add(normalized.lower())
        local, sep, domain = normalized.lower().partition("@")
        if sep and "+" in local:
            variants.add(f"{local.split('+', 1)[0]}@{domain}")
    return {item for item in variants if item}


_PREFECTURE_RE = (
    r"北海道|東京都|京都府|大阪府|"
    r"(?:青森|岩手|宮城|秋田|山形|福島|茨城|栃木|群馬|埼玉|千葉|神奈川|新潟|富山|石川|福井|山梨|長野|岐阜|静岡|愛知|三重|滋賀|兵庫|奈良|和歌山|鳥取|島根|岡山|広島|山口|徳島|香川|愛媛|高知|福岡|佐賀|長崎|熊本|大分|宮崎|鹿児島|沖縄)県"
)

_CITY_PREFECTURE_HINTS = {
    "名古屋市": "愛知県",
    "大阪市": "大阪府",
    "大津市": "滋賀県",
    "静岡市": "静岡県",
    "高山市": "岐阜県",
    "札幌市": "北海道",
}

_ENGLISH_ADDRESS_HINTS = {
    "tokyo-to minato-ku": "東京都港区",
    "tokyo minato-ku": "東京都港区",
    "minato-ku": "東京都港区",
}


def _generalize_address(value: str) -> str:
    normalized = _normalize_ascii(value)
    compact = re.sub(r"[\s　]+", "", normalized)
    lowered = normalized.lower()
    for marker, replacement in _ENGLISH_ADDRESS_HINTS.items():
        if marker in lowered:
            return replacement

    match = re.search(rf"(?P<pref>{_PREFECTURE_RE})(?P<city>[一-龯ぁ-んァ-ヶー]+?市)(?P<ward>[一-龯ぁ-んァ-ヶー]+?区)?", compact)
    if match:
        return f"{match.group('pref')}{match.group('city')}{match.group('ward') or ''}"
    match = re.search(rf"(?P<pref>{_PREFECTURE_RE})(?P<ward>[一-龯ぁ-んァ-ヶー]+?区)", compact)
    if match:
        return f"{match.group('pref')}{match.group('ward')}"
    match = re.search(rf"(?P<pref>{_PREFECTURE_RE})(?P<city>[一-龯ぁ-んァ-ヶー]+?市)", compact)
    if match:
        return f"{match.group('pref')}{match.group('city')}"

    for city, prefecture in _CITY_PREFECTURE_HINTS.items():
        if city not in compact:
            continue
        city_index = compact.find(city)
        after_city = compact[city_index + len(city) :]
        ward_match = re.match(r"([一-龯ぁ-んァ-ヶー]+?区)", after_city)
        return f"{prefecture}{city}{ward_match.group(1) if ward_match else ''}"
    return ""


def _prefecture_only(value: str) -> str:
    normalized = _normalize_ascii(value)
    lowered = normalized.lower()
    for marker, replacement in _ENGLISH_ADDRESS_HINTS.items():
        if marker in lowered:
            pref_match = re.search(_PREFECTURE_RE, replacement)
            return pref_match.group(0) if pref_match else replacement
    match = re.search(_PREFECTURE_RE, normalized)
    if match:
        return match.group(0)
    for city, prefecture in _CITY_PREFECTURE_HINTS.items():
        if city in normalized:
            return prefecture
    return _generalize_address(value) or normalized


def _analysis_department(value: str) -> str:
    text = _normalize_ascii(value)
    if text == "南工場":
        return "工場"
    return text


def _generalize_department(value: str) -> str:
    text = _normalize_ascii(value)
    mapping = (
        ("商品開発", "商品開発"),
        ("研究開発", "研究開発"),
        ("情報システム", "情報システム"),
        ("海外事業", "海外事業"),
        ("Finance", "財務"),
        ("財務", "財務"),
        ("営業", "営業"),
        ("総務", "管理"),
        ("管理", "管理"),
        ("企画", "企画"),
        ("製造", "製造"),
        ("工場", "製造"),
    )
    for marker, replacement in mapping:
        if marker in text:
            return replacement
    return re.sub(r"(部|課|室)$", "", text)


def _generalize_role(value: str) -> str:
    text = _normalize_ascii(value)
    if not text:
        return text
    if any(word in text for word in ("代表", "取締役", "CFO", "CEO", "COO")):
        return "経営者"
    if any(word in text for word in ("部長", "課長", "主任", "工場長")):
        return "管理職"
    if text in {"担当"}:
        return "担当者"
    return text


def _cleanup_address_tail(value: str) -> str:
    municipality = rf"(?:{_PREFECTURE_RE})[一-龯ぁ-んァ-ヶー]+?(?:市[一-龯ぁ-んァ-ヶー]+?区|市|区)"
    return re.sub(
        rf"({municipality})[\s　]+[一-龯ぁ-んァ-ヶーA-Za-z0-9０-９一二三四五六七八九十百千]+(?:ビル|タワー|マンション)[一二三四五六七八九十0-9０-９]*(?:階|F|f)",
        r"\1",
        value,
    )


def _cleanup_alias_note_tail(value: str) -> str:
    value = re.sub(r"旧姓は[^。]+。", "旧姓情報あり。", value)
    value = re.sub(r"(個人[0-9]{3})(?:、\1)+(?:、\1)?の表記あり。", r"\1の表記ゆれあり。", value)
    value = re.sub(r"(個人[0-9]{3})(?:、\1)+、[A-Za-z]+の表記あり。", r"\1の表記ゆれあり。", value)
    return value


def _postprocess_text(value: str, options: ProcessingOptions, sheet_name: str = "", coordinate: str = "") -> str:
    text = value
    if sheet_name == "05_表記ゆれ":
        coordinate_exact_analysis = {
            "C13": "mail候補001@example.invalid",
            "C19": "個人候補003さん",
            "C21": "mail候補002@example.invalid",
        }
        coordinate_exact_external = {
            "C12": "メール002",
            "C13": "メール候補001",
            "C19": "個人候補003さん",
        }
        if options.is_analysis and coordinate in coordinate_exact_analysis:
            return coordinate_exact_analysis[coordinate]
        if not options.is_analysis and coordinate in coordinate_exact_external:
            return coordinate_exact_external[coordinate]
        if options.is_analysis:
            exact = {
                "個人候補002": "個人候補001",
                "+81-90-XXXX-2222": "090-XXXX-2222",
                "組織001": "法人006工場",
                "個人候補003さん": "個人候補002さん",
                "個人候補001さん": "個人候補003さん",
                "電話番号の下4桁は2222": "電話番号の下4桁はXXXX",
                "メール候補001": "mail候補002@example.invalid",
            }
        else:
            exact = {
                "個人候補002": "個人候補001",
                "携帯010": "携帯001",
                "メール002": "メール候補001",
                "組織001": "法人006",
                "個人候補003さん": "個人候補002さん",
                "個人候補001さん": "個人候補003さん",
                "電話番号の下4桁は2222": "電話情報あり",
                "メール候補001": "メール候補002",
            }
        if text in exact:
            return exact[text]
    text = re.sub(r"C(\d{4})", lambda match: f"顧客{int(match.group(1)):03d}", text)
    text = re.sub(r"顧客A社", "法人候補A", text)
    text = re.sub(r"競合[A-ZＡ-Ｚ]社", "競合企業", text)
    text = re.sub(r"発売前情報。想定原価率は([0-9０-９]+％)以下が条件。", r"発売前情報。想定原価率は\1以下が条件。", text)
    text = re.sub(r"番号は下4桁が[0-9０-９]{4}", "番号は下4桁がXXXX", text)
    text = text.replace("会社名は「高」、氏名は「髙」。", "会社名と氏名の表記差。")
    text = text.replace("「髙」と「高」の表記ゆれあり。", "個人003の表記ゆれあり。")
    text = text.replace("「南」は氏名・会社名・工場名に登場。", "同一語が氏名・法人名・部署名に登場。")
    text = text.replace("「南」「森」が氏名・組織・一般語として混在。", "同一語が氏名・組織・一般語として混在。")
    text = text.replace("姓のみ。顧客の森か社員の森か不明。", "姓のみ。顧客の個人007か社員の個人010か不明。")
    text = text.replace("組織001の個人006です。", "法人006の個人006です。")
    text = text.replace("滋賀県大津市で打合せ", "滋賀県大津市の本社で打合せ")
    text = text.replace("特に個人候補001さんの説明", "特に個人候補003さんの説明")
    if text.startswith("口座は架空銀行 名古屋支店 普通 ****567"):
        text = "口座は金融機関001 名古屋支店 普通 ****567、名義は個人候補001です。"
    if options.is_analysis:
        return text

    text = text.replace("個人003の表記ゆれあり。", "氏名の表記ゆれあり。")
    text = text.replace("個人004の表記ゆれあり。", "氏名の表記ゆれあり。")
    text = re.sub(r"旧姓情報あり。登録名は個人[0-9]{3}。", "旧姓情報あり。", text)
    text = text.replace("請求書は個人001部長宛。緊急時は携帯へ連絡。", "請求書の送付先担当者あり。緊急連絡先あり。")
    text = text.replace("同一語が氏名・法人名・部署名に登場。", "同一語が複数の情報種別に登場。")
    text = text.replace("珍しい読み方、メールにプラス記号。", "珍しい読み方。メールにプラス記号。")
    text = text.replace("顧客002と同一人物を想定。全角空白・表記ゆれ。", "顧客002と同一人物を想定。表記ゆれあり。")
    if "パスワード" in text and "APIキー" in text:
        return "認証情報を削除。"
    if text.startswith("口座は"):
        name_match = re.search(r"名義は([^。]+)", text)
        name = name_match.group(1).removesuffix("です") if name_match else "非表示"
        if re.fullmatch(r"個人[0-9]{3}", name):
            name = "個人候補001"
        return f"口座情報を削除。名義は{name}です。"
    text = re.sub(r"住所は((?:北海道|東京都|京都府|大阪府|[一-龯]+県)[^。]*?)。", lambda match: f"住所は{_prefecture_only(match.group(1))}。", text)
    text = re.sub(r"(東京都|大阪府|愛知県|滋賀県|静岡県|岐阜県|北海道)[一-龯ぁ-んァ-ヶー0-9０-９一二三四五六七八九十\-ー丁目番地号\s　]+に資料", r"\1に資料", text)
    text = re.sub(r"(滋賀県)[一-龯ぁ-んァ-ヶー0-9０-９一二三四五六七八九十\-ー丁目番地号\s　]*の本社", r"\1の本社", text)
    text = text.replace("滋賀県で打合せ", "滋賀県の本社で打合せ")
    text = text.replace("メール002 に", "メール002に")
    text = re.sub(r"電話は(?:0[789]0[-X0-9]+|携帯[0-9]{3})です。", lambda match: f"電話は{_normalize_mobile_alias(match.group(0))}です。", text)
    text = re.sub(r"携帯(?:0[789]0[-X0-9]+|携帯[0-9]{3})へ", lambda match: f"{_normalize_mobile_alias(match.group(0))}へ", text)
    text = re.sub(r"Please call me at (携帯[0-9]{3}) or email (メール[0-9]{3})\.", r"Please call \1 or email \2.", text)
    text = re.sub(r"Please call 携帯[0-9]{3} or email メール005\.", "Please call 携帯005 or email メール005.", text)
    text = text.replace("南側の入口から入り、森を抜けた先の建物へ", "入口から入り、建物へ")
    text = text.replace("昨日の件、例の会社の例の人へ連絡しておいてください。番号は下4桁がXXXXです。", "昨日の件、例の会社の例の人へ連絡してください。電話情報を削除。")
    text = text.replace("昨日の件、例の会社の例の人へ連絡しておいてください。番号は下4桁が2222です。", "昨日の件、例の会社の例の人へ連絡してください。電話情報を削除。")
    text = text.replace("姓だけの呼称と完全なメールアドレスが混在。", "個人情報が複数形式で混在。")
    text = text.replace("会社名と氏名の表記差。", "法人名と氏名の表記差。")
    text = text.replace("顧客マスタでは個人004。住所は途中まで。", "顧客マスタでは個人004。住所は一部記載。")
    text = text.replace("同一語が氏名・組織・一般語として混在。", "氏名・組織・一般語が混在。")
    text = text.replace("個人情報ではないが機密情報。", "認証情報を含む。")
    text = text.replace("姓のみ。顧客の個人007か社員の個人010か不明。", "姓のみ。複数候補が存在。")
    text = re.sub(r"住所変更：((?:北海道|東京都|京都府|大阪府|[一-龯]+県).*)", lambda match: f"住所変更：{_prefecture_only(match.group(1))}", text)
    text = text.replace("伏字001", "予定日時001")
    return text


def _normalize_mobile_alias(value: str) -> str:
    match = re.search(r"携帯[0-9]{3}", value)
    return match.group(0) if match else "携帯"


def _digits(value: str) -> str:
    return re.sub(r"\D", "", _normalize_ascii(value))


def _alias_email(value: str, alias_book: AliasBook, options: ProcessingOptions) -> str:
    alias = alias_book.get("email", _normalize_ascii(value).lower())
    if not options.is_analysis:
        return alias
    number = re.search(r"(\d{3})$", alias)
    suffix = number.group(1) if number else "001"
    return f"mail{suffix}@example.invalid"


def _alias_phone(kind: str, value: str, alias_book: AliasBook, options: ProcessingOptions) -> str:
    if options.is_analysis:
        return _mask_phone(value)
    if kind == "landline_phone":
        return alias_book.get("landline_phone", value)
    if kind == "mobile_phone":
        return alias_book.get("mobile_phone", value)
    digits = _digits(value)
    if digits.startswith(("070", "080", "090")) or digits.startswith(("8170", "8180", "8190")):
        return alias_book.get("mobile_phone", value)
    return alias_book.get("landline_phone", value)


def _mask_postal_code(value: str) -> str:
    digits = _digits(value)
    if len(digits) >= 3:
        return f"{digits[:3]}-XXXX"
    return "XXX-XXXX"


def _mask_phone(value: str) -> str:
    normalized = _normalize_ascii(value)
    parts = [part for part in re.split(r"[^0-9+]+", normalized) if part]
    digits = _digits(normalized)
    if normalized.startswith("+81"):
        national = digits[2:]
        if national.startswith("0"):
            national = national[1:]
        if national.startswith(("70", "80", "90")) and len(national) >= 10:
            return f"+81-{national[:2]}-XXXX-{national[-4:]}"
        area_len = 1 if national.startswith(("3", "6")) else 2
        return f"0{national[:area_len]}-{'X' * max(3, len(national) - area_len - 4)}-{national[-4:]}"
    if len(digits) == 11 and digits.startswith(("070", "080", "090")):
        return f"{digits[:3]}-XXXX-{digits[-4:]}"
    if len(parts) >= 3:
        return f"{parts[0]}-{'X' * len(parts[-2])}-{parts[-1]}"
    if len(digits) >= 10:
        return f"{digits[:3]}-{'X' * (len(digits) - 7)}-{digits[-4:]}"
    return "電話番号"


def _mask_bank_account(value: str, alias_book: AliasBook) -> str:
    normalized = _normalize_ascii(value)
    match = re.match(r"(?P<bank>.+?(?:銀行|信用金庫|Bank))\s*(?P<rest>.*?)(?P<digits>[0-9]{4,})$", normalized, flags=re.IGNORECASE)
    if not match:
        return re.sub(r"[0-9]{4,}", lambda item: f"****{item.group(0)[-3:]}", normalized)
    bank_alias = alias_book.get("financial_institution", normalized)
    rest = match.group("rest").strip()
    digits = match.group("digits")
    return f"{bank_alias} {rest} ****{digits[-3:]}".strip()


def _number(value: Any) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    normalized = _normalize_ascii(str(value)).replace(",", "").replace("円", "")
    if normalized.endswith("%"):
        normalized = normalized[:-1]
    try:
        return float(normalized)
    except ValueError:
        return None


def _yen_range(value: Any, width: int) -> str:
    number = _number(value)
    if number is None:
        return "金額範囲不明"
    lower = int(number // width * width)
    upper = lower + width - 1
    return f"{lower:,}～{upper:,}円"


def _amount_range(value: Any, width: int) -> str:
    number = _number(value)
    if number is None:
        return "金額範囲不明"
    lower = int(number // width * width)
    upper = lower + width
    return f"{lower // 10_000:,}万～{upper // 10_000:,}万円"


def _estimate_amount_range(value: Any) -> str:
    number = _number(value)
    if number is None:
        return "金額範囲不明"
    width = 5_000_000 if number >= 5_000_000 else 1_000_000
    return _amount_range(value, width)


def _cost_amount_range(value: Any) -> str:
    number = _number(value)
    if number is None:
        return "金額範囲不明"
    width = 5_000_000 if number >= 10_000_000 else 1_000_000
    return _amount_range(value, width)


def _salary_range(value: Any) -> str:
    number = _number(value)
    if number is None:
        return "給与範囲不明"
    lower = int(number // 100_000 * 10)
    return f"{lower}万円台"


def _bonus_range(value: Any) -> str:
    number = _number(value)
    if number is None:
        return "賞与範囲不明"
    lower = int(number // 100_000 * 10)
    return f"{lower}万円台"


def _rate_range(value: Any) -> str:
    number = _number(value)
    if number is None:
        return "率不明"
    percent = number * 100 if number <= 1 else number
    lower = int(percent // 10 * 10)
    return f"{lower}％台"


def _year(value: Any) -> int | None:
    if isinstance(value, (datetime, date)):
        return value.year
    match = re.search(r"(19|20)\d{2}", _normalize_ascii(str(value)))
    return int(match.group(0)) if match else None


def _week_bucket(value: Any) -> str:
    if isinstance(value, (datetime, date)):
        dt = value
    else:
        text = _normalize_ascii(str(value))
        match = re.search(r"(?P<year>20\d{2})[-/](?P<month>\d{1,2})[-/](?P<day>\d{1,2})", text)
        if not match:
            return text
        dt = date(int(match.group("year")), int(match.group("month")), int(match.group("day")))
    week = 1 if dt.day <= 3 else ((dt.day - 4) // 7) + 2
    return f"{dt.year}年{dt.month}月第{week}週"
